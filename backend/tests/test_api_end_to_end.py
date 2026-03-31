import tempfile
import unittest
import uuid
from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import load_workbook

import backend.tests.warnings_config  # noqa: F401
from backend.main import app
from backend.tests.test_support import ensure_database_schema


class ApiEndToEndTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        ensure_database_schema()

    def setUp(self):
        self.client = TestClient(app)
        email = f"e2e_{uuid.uuid4().hex[:8]}@testmail.com"
        register = self.client.post(
            '/api/auth/register/',
            json={'name': 'E2E Tester', 'email': email, 'password': 'StrongPass123!'},
        )
        self.assertEqual(register.status_code, 200, msg=register.text)

    def tearDown(self):
        self.client.close()

    def test_full_user_flow(self):
        upload = self.client.post(
            '/api/documents/',
            files={'file': ('notes.txt', b'Agent integration document', 'text/plain')},
        )
        self.assertEqual(upload.status_code, 200, msg=upload.text)
        document = upload.json()

        db_conn = self.client.post(
            '/api/database/connection/',
            json={'mode': 'sqlite', 'displayName': 'E2E DB', 'sqlitePath': './backend/axon.db'},
        )
        self.assertEqual(db_conn.status_code, 200, msg=db_conn.text)

        chat = self.client.post(
            '/api/chat/',
            json={
                'message': 'Use the attached document and give me a short summary.',
                'document_ids': [str(document['id'])],
            },
        )
        self.assertEqual(chat.status_code, 200, msg=chat.text)
        conversation = chat.json()
        conversation_id = conversation['id']

        self.assertTrue(conversation.get('messages'))
        assistant_messages = [m for m in conversation['messages'] if m.get('sender') == 'assistant']
        self.assertTrue(assistant_messages)
        self.assertTrue(assistant_messages[-1]['content'].strip())

        export_docx = self.client.get(f'/api/conversations/{conversation_id}/export/')
        self.assertEqual(export_docx.status_code, 200, msg=export_docx.text)

        export_zip = self.client.post(
            f'/api/conversations/{conversation_id}/export/zip/',
            json={'sqlResults': []},
        )
        self.assertEqual(export_zip.status_code, 200, msg=export_zip.text)

        feedback_target = assistant_messages[-1]['id']
        feedback_add = self.client.post(
            f'/api/messages/{feedback_target}/feedback/',
            json={'type': 'like', 'reason': 'works'},
        )
        self.assertEqual(feedback_add.status_code, 200, msg=feedback_add.text)

        feedback_delete = self.client.delete(f'/api/messages/{feedback_target}/feedback/')
        self.assertEqual(feedback_delete.status_code, 200, msg=feedback_delete.text)

        prefs_update = self.client.put('/api/preferences/', json={'theme': 'light'})
        self.assertEqual(prefs_update.status_code, 200, msg=prefs_update.text)

        db_upload_payload = None
        with tempfile.NamedTemporaryFile(suffix='.db') as db_file:
            db_upload = self.client.post(
                '/api/database/upload/',
                files={'database': ('temp.db', db_file, 'application/octet-stream')},
            )
            self.assertEqual(db_upload.status_code, 200, msg=db_upload.text)
            db_upload_payload = db_upload.json()

        self.assertTrue(db_upload_payload['path'])

        export_xlsx = self.client.post(
            '/api/database/export/',
            json={
                'query': 'SELECT 1 AS id, "Alice" AS name;',
                'columns': ['id', 'name'],
                'rows': [{'id': 1, 'name': 'Alice'}],
            },
        )
        self.assertEqual(export_xlsx.status_code, 200, msg=export_xlsx.text)
        self.assertIn(
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            export_xlsx.headers.get('content-type', ''),
        )

        workbook = load_workbook(BytesIO(export_xlsx.content))
        sheet = workbook.active
        self.assertEqual(sheet.cell(1, 1).value, 'id')
        self.assertEqual(sheet.cell(1, 2).value, 'name')
        self.assertEqual(sheet.cell(2, 1).value, 1)
        self.assertEqual(sheet.cell(2, 2).value, 'Alice')
        workbook.close()

        delete_conversation = self.client.delete(f'/api/conversations/{conversation_id}/?delete_files=true')
        self.assertEqual(delete_conversation.status_code, 200, msg=delete_conversation.text)


if __name__ == '__main__':
    unittest.main()
